"""Mengisi template HRD dengan pegawai yang SUDAH ada, supaya HRD tinggal
melengkapi kolom yang kosong.

Berkas kosongnya TIDAK dibangun ulang dari nol — yang dilakukan di sini adalah
membuka hasil `_buat-template-pegawai.py` lalu menulis barisnya. Dengan begitu
dropdown, rumus kolom Cek, lebar kolom, dan seluruh gayanya tetap utuh; menyalin
ulang semua itu berarti dua tempat yang harus dijaga sejalan, dan suatu hari
hanya satu yang diperbarui.

Label diambil dari `src/lib/constants.ts`, BUKAN diketik ulang di sini. Nama
yang berbeda satu huruf pun akan ditolak importer, dan HRD-lah yang menanggung
kebingungannya.

Pakai:
    python3 scripts/_isi-template-pegawai.py <pegawai.json> [keluaran.xlsx]

<pegawai.json> berisi array objek dengan kunci: employeeNo, fullName, jobTitle,
jobLevel, employeeType, workPattern, joined, kmulai, kakhir, address, atasan,
email, isActive, divisi, birthPlace, lahir, education, bloodType.
"""
import json
import os
import re
import sys
from datetime import date

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

AKAR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..")
KOSONG = os.path.join(AKAR, "docs", "template", "Template-Data-Pegawai.xlsx")
KONSTANTA = os.path.join(AKAR, "src", "lib", "constants.ts")

FONT = "Arial"
ISI_FILL = PatternFill("solid", fgColor="FFF2CC")   # kolom yang menunggu diisi
HDR_ROW, FIRST = 3, 4


def pasangan(nama: str) -> dict[str, str]:
    """Membaca satu konstanta pasangan [kode, label] dari constants.ts.

    Dibaca dari berkasnya, bukan disalin: importer mencocokkan label per huruf.
    """
    teks = open(KONSTANTA, encoding="utf-8").read()
    blok = re.search(rf"export const {nama} = \[(.*?)\] as const;", teks, re.S)
    if not blok:
        raise SystemExit(f"konstanta {nama} tidak ketemu di {KONSTANTA}")
    return {k: v for k, v in re.findall(r'\["([^"]+)",\s*"([^"]+)"\]', blok.group(1))}


def main() -> None:
    if len(sys.argv) < 2:
        raise SystemExit(__doc__)
    sumber = sys.argv[1]
    keluar = sys.argv[2] if len(sys.argv) > 2 else os.path.join(
        AKAR, "docs", "template", "Template-Data-Pegawai-TERISI.xlsx")

    orang = json.load(open(sumber, encoding="utf-8"))
    JENJANG, JENIS, POLA = pasangan("JOB_LEVELS"), pasangan("EMPLOYEE_TYPES"), pasangan("WORK_PATTERNS")
    DIDIK, DARAH = pasangan("EDUCATION_LEVELS"), pasangan("BLOOD_TYPES")

    wb = load_workbook(KOSONG)
    ws = wb["Data Pegawai"]

    # Judul dipetakan lewat NAMANYA, sama seperti importer. Kolom yang bergeser
    # tidak akan menaruh nilai di tempat yang salah.
    kolom = {
        str(ws.cell(row=HDR_ROW, column=i).value or "").replace("*", "").strip(): i
        for i in range(1, ws.max_column + 1)
    }

    def tanggal(s):
        return date.fromisoformat(s) if s else ""

    for n, o in enumerate(orang):
        r = FIRST + n
        nilai = {
            "NIK": o["employeeNo"],
            "Nama Lengkap": o["fullName"],
            "Jabatan": o["jobTitle"] or "",
            "Jenjang Jabatan": JENJANG.get(o["jobLevel"], o["jobLevel"]),
            "Status Kepegawaian": JENIS.get(o["employeeType"], o["employeeType"]),
            "Pola Kerja": POLA.get(o["workPattern"], o["workPattern"]),
            "Tanggal Bergabung": tanggal(o["joined"]),
            "Kontrak Mulai": tanggal(o["kmulai"]),
            "Kontrak Berakhir": tanggal(o["kakhir"]),
            "Alamat": o["address"] or "",
            "NIK Atasan": o["atasan"] or "",
            "Email Akun CRM": o["email"] or "",
            "Aktif": "Ya" if o["isActive"] else "Tidak",
            "Divisi": o["divisi"] or "",
            # Empat kolom yang MENUNGGU DIISI. Yang sudah terisi dipertahankan
            # supaya menjalankan skrip ini dua kali tidak menghapus kerja HRD.
            "Tempat Lahir": o.get("birthPlace") or "",
            "Tanggal Lahir": tanggal(o.get("lahir")),
            "Pendidikan Terakhir": DIDIK.get(o.get("education") or "", ""),
            "Golongan Darah": DARAH.get(o.get("bloodType") or "", ""),
        }
        for judul, isi in nilai.items():
            c = ws.cell(row=r, column=kolom[judul], value=isi)
            c.font = Font(name=FONT, size=10)
            if isinstance(isi, date):
                c.number_format = "yyyy-mm-dd"
        # Empat kolom itu diberi warna supaya HRD langsung tahu mana tugasnya.
        for judul in ("Tempat Lahir", "Tanggal Lahir", "Pendidikan Terakhir", "Golongan Darah"):
            ws.cell(row=r, column=kolom[judul]).fill = ISI_FILL

    ws["A2"] = (
        f"{len(orang)} pegawai yang sudah terdaftar sudah terisi - JANGAN diubah. "
        "Yang perlu diisi HRD hanya empat kolom berlatar kuning di kanan: "
        "Tempat Lahir, Tanggal Lahir, Pendidikan Terakhir, Golongan Darah. "
        "Golongan darah WAJIB pakai tanda + atau -; yang belum tahu pilih Tidak diketahui. "
        "Pegawai baru boleh ditambahkan di baris kosong di bawahnya."
    )
    ws["A2"].font = Font(name=FONT, size=9, italic=True, color="C00000")

    wb.save(keluar)
    print(f"tersimpan: {keluar} ({len(orang)} pegawai terisi)")


if __name__ == "__main__":
    main()
