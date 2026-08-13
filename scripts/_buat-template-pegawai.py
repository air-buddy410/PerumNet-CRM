"""Template Excel input data pegawai PerumNet CRM.

Nilai dropdown diambil dari konstanta yang benar-benar dipakai kode
(EMPLOYEE_TYPES, WORK_PATTERNS, JOB_LEVELS di src/lib/constants.ts), memakai
LABEL Indonesia supaya HRD tidak perlu tahu kode internalnya. Pemetaan kembali
ke kode dilakukan saat impor.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUT = "/Users/air_buddy/Dev Project/PerumNet-CRM/docs/template/Template-Data-Pegawai.xlsx"

FONT = "Arial"
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
REQ_FILL = PatternFill("solid", fgColor="FFF2CC")   # kolom wajib
EX_FILL = PatternFill("solid", fgColor="EAF3EA")    # baris contoh
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# (header, lebar, wajib?, catatan singkat)
COLS = [
    ("NIK",                 14, False, "KOSONGKAN saja - sistem menerbitkannya otomatis (10000001, 10000002, ...). Isi hanya bila orang ini sudah punya NIK lama."),
    ("Nama Lengkap",        26, True,  ""),
    ("Jabatan",             22, False, "Teks bebas, mis. Teknisi Lapangan"),
    ("Jenjang Jabatan",     16, True,  "Staff atau Leader"),
    ("Status Kepegawaian",  20, True,  "Tetap / Paruh Waktu / Kontrak / Masa Percobaan"),
    ("Pola Kerja",          13, True,  "Shift atau Non-Shift"),
    ("Tanggal Bergabung",   18, True,  "Format YYYY-MM-DD"),
    ("Kontrak Mulai",       15, False, "HANYA untuk Status = Kontrak"),
    ("Kontrak Berakhir",    17, False, "WAJIB bila Status = Kontrak"),
    ("Alamat",              34, False, "Alamat domisili"),
    ("NIK Atasan",          14, False, "NIK atasan langsung, harus ada di daftar ini"),
    ("Email Akun CRM",      26, False, "Kosongkan bila belum punya akun sistem"),
    ("Aktif",               9,  True,  "Ya atau Tidak"),
    ("Divisi",              20, True,  "Divisi tempat orang ini bekerja"),
    ("Cek",                 40, False, "Terisi otomatis - jangan diketik"),
]

# Nama divisi HARUS sama persis dengan tabel Division di database. Kalau daftar
# di sana berubah, daftar ini ikut diperbarui — impor mencocokkannya per huruf,
# dan nama yang tidak dikenal DITOLAK, bukan dikosongkan diam-diam.
DIVISI = [
    "Customer Service", "Finance", "IT/DevOps", "Management", "Marketing",
    "NOC", "Operational", "Project", "Sales", "Warehouse",
]

HDR_ROW = 3
EX_ROW = 4
FIRST = 5
LAST = 204

wb = Workbook()
ws = wb.active
ws.title = "Data Pegawai"

# ── Judul & legenda ─────────────────────────────────────────────
ws["A1"] = "Data Pegawai PerumNet"
ws["A1"].font = Font(name=FONT, bold=True, size=14, color="1F4E79")
ws["A2"] = (
    "Isi mulai baris 5. Kolom berlatar kuning WAJIB diisi. "
    "Kolom NIK boleh dikosongkan - sistem yang menerbitkannya. "
    "Baris 4 adalah contoh - hapus sebelum dikirim. "
    "Kolom Cek terisi sendiri; perbaiki dulu semua yang bertanda PERIKSA."
)
ws["A2"].font = Font(name=FONT, size=9, italic=True, color="595959")
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(COLS))

# ── Header ──────────────────────────────────────────────────────
for i, (name, width, required, note) in enumerate(COLS, start=1):
    c = ws.cell(row=HDR_ROW, column=i, value=name + (" *" if required else ""))
    c.font = HDR_FONT
    c.fill = HDR_FILL
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER
    ws.column_dimensions[get_column_letter(i)].width = width
    if note:
        c.comment = None  # catatan lengkap ada di sheet Panduan
ws.row_dimensions[HDR_ROW].height = 30

# ── Baris contoh ────────────────────────────────────────────────
EXAMPLE = [
    "", "Teguh Santoso", "Teknisi Lapangan", "Staff", "Kontrak",
    "Shift", "2026-01-06", "2026-01-06", "2026-12-31",
    "Jl. Melati No. 7, Denpasar", "", "teguh@perumnet.id", "Ya", "Operational",
]
for i, val in enumerate(EXAMPLE, start=1):
    c = ws.cell(row=EX_ROW, column=i, value=val)
    c.font = Font(name=FONT, size=10, italic=True, color="375623")
    c.fill = EX_FILL
    c.border = BORDER

# ── Sel data kosong ─────────────────────────────────────────────
for r in range(FIRST, LAST + 1):
    for i, (_, _, required, _) in enumerate(COLS, start=1):
        c = ws.cell(row=r, column=i)
        c.font = Font(name=FONT, size=10)
        c.border = BORDER
        if required and i != len(COLS):
            c.fill = REQ_FILL

# Kolom tanggal ditampilkan sebagai tanggal.
for col in ("G", "H", "I"):
    for r in range(EX_ROW, LAST + 1):
        ws[f"{col}{r}"].number_format = "yyyy-mm-dd"

# ── Dropdown ────────────────────────────────────────────────────
def dropdown(col: str, options: list[str], title: str, msg: str) -> None:
    dv = DataValidation(
        type="list",
        formula1='"' + ",".join(options) + '"',
        allow_blank=True,
        showDropDown=False,   # False = panah dropdown DITAMPILKAN (kuirk openpyxl)
    )
    dv.error = msg
    dv.errorTitle = title
    dv.showErrorMessage = True
    ws.add_data_validation(dv)
    dv.add(f"{col}{EX_ROW}:{col}{LAST}")

dropdown("D", ["Staff", "Leader"], "Jenjang tidak dikenal",
         "Pilih Staff atau Leader dari daftar.")
dropdown("E", ["Karyawan Tetap", "Paruh Waktu", "Kontrak", "Masa Percobaan"],
         "Status tidak dikenal", "Pilih salah satu dari daftar.")
dropdown("F", ["Non-Shift", "Shift"], "Pola kerja tidak dikenal",
         "Pilih Non-Shift atau Shift dari daftar.")
dropdown("M", ["Ya", "Tidak"], "Isian tidak dikenal", "Pilih Ya atau Tidak.")
dropdown("N", DIVISI, "Divisi tidak dikenal",
         "Pilih divisi dari daftar. Nama di luar daftar akan ditolak saat impor.")

# ── Kolom Cek ───────────────────────────────────────────────────
# Menangkap tiga kesalahan yang paling sering terjadi dan paling menyusahkan
# saat impor: kolom wajib kosong, Kontrak tanpa tanggal berakhir, dan tanggal
# kontrak terisi pada jenis yang bukan Kontrak (tanggal nyasar itu yang
# dipakai penyapu untuk membekukan akun).
for r in range(EX_ROW, LAST + 1):
    ws[f"O{r}"] = (
        f'=IF(B{r}="","",'
        f'IF(OR(D{r}="",E{r}="",F{r}="",G{r}="",M{r}="",N{r}=""),'
        f'"PERIKSA: ada kolom wajib yang kosong",'
        f'IF(AND(E{r}="Kontrak",I{r}=""),'
        f'"PERIKSA: Kontrak wajib punya Kontrak Berakhir",'
        f'IF(AND(E{r}<>"Kontrak",OR(H{r}<>"",I{r}<>"")),'
        f'"PERIKSA: hanya Kontrak yang boleh diisi tanggal kontrak",'
        f'IF(AND(I{r}<>"",H{r}<>"",I{r}<=H{r}),'
        f'"PERIKSA: Kontrak Berakhir harus setelah Kontrak Mulai",'
        f'"OK")))))'
    )
    ws[f"O{r}"].font = Font(name=FONT, size=9, color="C00000")
    ws[f"O{r}"].border = BORDER

ws.freeze_panes = "A4"

# ── Sheet Panduan ───────────────────────────────────────────────
p = wb.create_sheet("Panduan")
p["A1"] = "Panduan Pengisian"
p["A1"].font = Font(name=FONT, bold=True, size=14, color="1F4E79")

guide = [
    ("", ""),
    ("Aturan umum", ""),
    ("Baris contoh", "Baris 4 berwarna hijau adalah CONTOH. Hapus sebelum file dikirim."),
    ("Kolom wajib", "Berlatar kuning dan bertanda * pada judulnya."),
    ("Kolom Cek", "Terisi otomatis. Kirim file hanya bila semua baris bertulis OK."),
    ("Urutan baris", "Bebas, tetapi atasan sebaiknya ditulis lebih dulu agar mudah diperiksa."),
    ("", ""),
    ("Penjelasan tiap kolom", ""),
]
for name, width, required, note in COLS:
    if note:
        guide.append((name + (" *" if required else ""), note))

guide += [
    ("", ""),
    ("Hal yang sering keliru", ""),
    ("Tanggal kontrak", "HANYA diisi bila Status Kepegawaian = Kontrak. Pada jenis lain harus KOSONG."),
    ("", "Alasannya: sistem membekukan akun otomatis saat tanggal kontrak lewat. Tanggal yang tertinggal pada karyawan tetap akan membekukan orang yang masih bekerja."),
    ("NIK Atasan", "Boleh diisi NIK, boleh juga NAMA LENGKAP persis seperti tertulis di kolom Nama Lengkap. Nama dipakai karena pada pengisian pertama belum ada seorang pun yang punya NIK."),
    ("", "Kalau ada DUA orang bernama sama di file ini, rujukan namanya ditolak dan diminta memakai NIK - bukan diambil salah satu."),
    ("Divisi", "Pilih dari dropdown. Ini yang menentukan pegawai masuk kelompok mana, dan dipakai untuk melabeli kotak email serta akses ke aplikasi PerumNet lainnya."),
    ("", "Divisi BUKAN hak akses. Peran dan kewenangan di CRM tetap ditetapkan IT, tidak bisa ditentukan dari file ini."),
    ("Email Akun CRM", "Hanya untuk yang akan punya akun sistem. Alamatnya harus sama persis dengan email di Authentik, kalau tidak yang bersangkutan tidak bisa login."),
    ("Format tanggal", "YYYY-MM-DD, contoh 2026-01-06."),
    ("NIK", "KOSONGKAN saja. Sistem menerbitkannya berurutan mulai 10000001 - delapan angka diawali 1, dan dijamin tidak kembar walau dua orang menyimpan bersamaan."),
    ("", "Isi manual HANYA bila orang itu sudah punya NIK lama yang menempel di dokumen lain. Nomor yang sudah dipakai otomatis dilewati sistem."),
    ("", "Kenapa diawali 1: dengan begitu tidak pernah ada nol di depan. Excel membuang nol di depan, dan kolom NIK Atasan justru tempat nomor itu diketik ulang - kalau berubah, impor tidak menemukan orangnya."),
]

for i, (a, b) in enumerate(guide, start=2):
    p[f"A{i}"] = a
    p[f"B{i}"] = b
    p[f"A{i}"].font = Font(name=FONT, size=10, bold=(b == "" and a != ""))
    p[f"B{i}"].font = Font(name=FONT, size=10)
    p[f"B{i}"].alignment = Alignment(wrap_text=True, vertical="top")
p.column_dimensions["A"].width = 24
p.column_dimensions["B"].width = 96

wb.save(OUT)
print("tersimpan:", OUT)
